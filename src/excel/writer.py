"""Excel workbook writing - writes values to specific cells"""
import warnings
from openpyxl import load_workbook


class ExcelWriter:
    """
    Writes values to Excel cells. Stateless - opens, writes, saves,
    and closes the workbook on each call.
    """

    def write_cell(self, file_path: str, sheet_name: str, row: int, column_name: str, value: str) -> bool:
        """
        Write a value to a specific cell identified by row number and column header name.

        Opens the workbook in write mode, finds the column by matching the header
        name in row 1, writes the value, saves, and closes.

        Args:
            file_path: Path to the .xlsx file
            sheet_name: Name of the worksheet
            row: Excel row number (1-indexed, where row 1 is the header)
            column_name: Header name of the target column
            value: String value to write

        Returns:
            True if write was successful, False on failure
        """
        if not column_name:
            warnings.warn("Column name cannot be empty. Skipping write.")
            return False

        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            # Find column index by header name (row 1)
            col_index = None
            for col in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col).value
                if header_value is not None and str(header_value).strip() == column_name:
                    col_index = col
                    break

            if col_index is None:
                warnings.warn(f"Column '{column_name}' not found in sheet '{sheet_name}'. Skipping write.")
                wb.close()
                return False

            # Write value
            ws.cell(row=row, column=col_index, value=value)

            # Save and close
            wb.save(file_path)
            wb.close()
            return True

        except Exception as e:
            warnings.warn(f"Failed to write to Excel: {e}")
            return False
