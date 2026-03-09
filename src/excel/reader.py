"""Excel workbook reading and data extraction with formatted cell values"""
import warnings
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Optional, Iterator, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelReader:
    """
    Reads Excel workbooks and extracts formatted cell values.
    Uses read_only mode for large files to reduce memory usage.
    """
    
    MAX_ROWS = 10000
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel reader.
        
        Args:
            file_path: Path to the .xlsx file
        """
        self.file_path = file_path
        self._workbook = None
        self._worksheet = None
        self._headers: List[str] = []
        self._row_count = 0
        self._sheet_names: List[str] = []
    
    def load_workbook(self) -> List[str]:
        """
        Load the workbook and return list of sheet names.
        
        Returns:
            List of sheet names in the workbook
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file is not a valid Excel file
        """
        try:
            # Load in read-only mode for large files
            self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            self._sheet_names = self._workbook.sheetnames
            return self._sheet_names
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")
    
    def select_sheet(self, sheet_name: str) -> None:
        """
        Select a worksheet and extract headers.
        
        Args:
            sheet_name: Name of the sheet to select
            
        Raises:
            ValueError: If sheet doesn't exist or is empty
        """
        if self._workbook is None:
            raise ValueError("Workbook not loaded. Call load_workbook() first.")
        
        if sheet_name not in self._sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        self._worksheet = self._workbook[sheet_name]
        
        # Extract headers from row 1 (1-indexed in openpyxl)
        self._headers = []
        header_row = next(self._worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        
        if header_row is None:
            raise ValueError(f"Sheet '{sheet_name}' is empty")
        
        for cell_value in header_row:
            if cell_value is not None:
                # Convert to string and strip whitespace
                header = str(cell_value).strip()
            else:
                # Use column letter for empty headers
                col_idx = len(self._headers) + 1
                header = get_column_letter(col_idx)
            self._headers.append(header)
        
        # Count data rows (excluding header)
        self._row_count = 0
        for _ in self._worksheet.iter_rows(min_row=2):
            self._row_count += 1
            if self._row_count > self.MAX_ROWS:
                warnings.warn(
                    f"Worksheet has more than {self.MAX_ROWS} rows. "
                    f"Only first {self.MAX_ROWS} rows will be processed."
                )
                break
    
    def get_headers(self) -> List[str]:
        """
        Get the list of column headers.
        
        Returns:
            List of header names
        """
        return self._headers.copy()
    
    def get_row_count(self) -> int:
        """
        Get the number of data rows (excluding header).
        
        Returns:
            Number of data rows
        """
        return self._row_count
    
    def get_sheet_names(self) -> List[str]:
        """
        Get the list of sheet names in the workbook.
        
        Returns:
            List of sheet names
        """
        return self._sheet_names.copy()
    
    def iterate_rows(
        self,
        start_row: int = 1,
        max_rows: Optional[int] = None
    ) -> Iterator[Dict[str, str]]:
        """
        Iterate over data rows and return formatted values as dictionaries.
        
        Args:
            start_row: 1-indexed row number to start from (1 = first data row)
            max_rows: Maximum number of rows to iterate (None = all rows)
            
        Yields:
            Dictionary mapping column names to formatted cell values
            
        Raises:
            ValueError: If no sheet is selected or start_row is invalid
        """
        if self._worksheet is None:
            raise ValueError("No sheet selected. Call select_sheet() first.")
        
        if start_row < 1 or start_row > self._row_count:
            raise ValueError(f"start_row must be between 1 and {self._row_count}")
        
        # Convert to 0-indexed for iteration
        start_row_idx = start_row
        
        # Track duplicate headers
        header_counts: Dict[str, int] = {}
        for header in self._headers:
            header_counts[header] = header_counts.get(header, 0) + 1
        
        duplicates = {h for h, c in header_counts.items() if c > 1}
        if duplicates:
            warnings.warn(
                f"Duplicate column headers detected: {', '.join(sorted(duplicates))}. "
                "First occurrence will be used."
            )
        
        rows_yielded = 0
        max_rows_limit = min(max_rows, self._row_count - start_row + 1) if max_rows else (self._row_count - start_row + 1)
        
        for row in self._worksheet.iter_rows(min_row=start_row + 1):
            if rows_yielded >= max_rows_limit:
                break
            
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(self._headers):
                    header = self._headers[col_idx]
                    # Skip duplicate headers (use first occurrence only)
                    if header in duplicates and self._headers.index(header) != col_idx:
                        continue
                    
                    row_data[header] = self._format_cell_value(cell)
            
            yield row_data
            rows_yielded += 1
    
    def _format_cell_value(self, cell) -> str:
        """
        Format a cell value to match what the user sees in Excel.
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            Formatted string value
        """
        if cell.value is None:
            return ""
        
        # Get the number format
        number_format = cell.number_format if cell.number_format else "General"
        
        # Handle dates
        if number_format and any(fmt in number_format.lower() for fmt in ['yy', 'mm', 'dd']):
            try:
                if isinstance(cell.value, datetime):
                    return cell.value.strftime(number_format.replace('mm', '%m').replace('dd', '%d').replace('yyyy', '%Y').replace('yy', '%y'))
            except:
                pass
        
        # Handle numbers with formatting
        if isinstance(cell.value, (int, float, Decimal)):
            # Simple formatting for common patterns
            if number_format == "General":
                return str(cell.value)
            elif "0.00" in number_format:
                return f"{float(cell.value):.2f}"
            elif "0.0" in number_format:
                return f"{float(cell.value):.1f}"
            elif "%" in number_format:
                return f"{float(cell.value) * 100:.2f}%"
            elif number_format.startswith("$"):
                return f"${float(cell.value):.2f}"
            else:
                return str(cell.value)
        
        # Handle booleans
        if isinstance(cell.value, bool):
            return "TRUE" if cell.value else "FALSE"
        
        # Default: convert to string
        return str(cell.value)
    
    def close(self) -> None:
        """Close the workbook and free resources"""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
            self._worksheet = None
    
    def __enter__(self):
        """Context manager entry"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        self.close()
        return False
